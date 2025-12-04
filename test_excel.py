from openpyxl import load_workbook
wb = load_workbook("services/api/templates/report_template.xlsx")
print(wb.sheetnames)
print(wb.active)
print(wb.active['A1'].value)
print(wb.active['B2'].value)
print(wb.active['C3'].value)
print(wb.active['D4'].value)
print(wb.active['E5'].value)
print(wb.active['F6'].value)
print(wb.active['G7'].value)
print(wb.active['H8'].value)

# Expected output:
# ['Report', 'Data', 'Summary']
# <Worksheet "Report">

# Report Title
# Student Name
# Student ID
# Date of Birth


            
