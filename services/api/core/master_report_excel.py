# services/api/core/master_report_excel.py
"""
Master Report Excel Generator

Generates a horizontal master inspection report showing:
- All master marks (rows)
- All inspection runs (columns)
- No PDF rendering, pure table-based report
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Template path
TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "master_report_template.xlsx"


class MasterReportRun:
    """Represents one inspection run (column in master report)"""
    
    def __init__(
        self,
        mark_set_id: str,
        mark_set_name: str,
        inspected_by: str,
        inspected_at: str,
        values: Dict[str, str],  # master_mark_id -> observed_value
    ):
        self.mark_set_id = mark_set_id
        self.mark_set_name = mark_set_name
        self.inspected_by = inspected_by
        self.inspected_at = inspected_at
        self.values = values


async def generate_master_report_excel(
    *,
    project_name: str,
    external_id: str,
    part_number: str,
    master_marks: List[Dict[str, Any]],
    runs: List[MasterReportRun],
    report_title: Optional[str] = None,
    max_runs: int = 100,
    dwg_num: Optional[str] = None,  # 🔥 NEW: optional drawing number for header
) -> bytes:
    """
    Generate master inspection report Excel.
    
    Args:
        project_name: Project name
        external_id: Business ID
        part_number: Part number
        master_marks: List of master mark dicts with:
            - mark_id (master_mark_id)
            - label
            - instrument (used as "Required Value")
            - is_required
            Plus optional: required_value, tol_min, tol_max
        runs: List of MasterReportRun objects (one per inspection map)
        report_title: Optional title override
        max_runs: Maximum runs to include (default 100)
    
    Returns:
        Excel file as bytes
    """
    
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Master report template not found: {TEMPLATE_PATH}")
    
    # Cap runs to prevent Excel explosion
    if len(runs) > max_runs:
        logger.warning(f"Capping runs from {len(runs)} to {max_runs}")
        runs = runs[:max_runs]
    
    # Load template
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    # --- 1) Fill header info ---
    id_str = f"{external_id} - {part_number}"
    # If dwg_num is provided, append it to the main header line
    if dwg_num:
        id_str = f"{id_str} (DWG: {dwg_num})"
    
    # ⚠️ IMPORTANT:
    # In the template, rows 3 and 4 are merged as:
    #   A3:G3  and  A4:G4
    # For merged ranges, ONLY the top-left cell (A3 / A4) is writable.
    # So we MUST write to A3 and A4, NOT E3/E4.
    ws["A3"] = id_str          # Full-width merged header line: "<external_id> - <part_number> (DWG: ...)"
    ws["A4"] = part_number     # Full-width merged header line: "<part_number>"


    
    # --- 2) Prepare styles from template ---
    # Template row 7 is first data row (after headers)
    template_row_idx = 7
    
    # Column E is the first "run" column in template
    first_run_col_idx = 5  # E = 5
    
    # Read template styles
    template_data_row_styles = {}
    for col_idx in range(1, 5):  # A-D (Label, Required, Tol Min, Tol Max)
        cell = ws.cell(row=template_row_idx, column=col_idx)
        template_data_row_styles[col_idx] = cell._style
    
    template_run_col_style = {}
    for row_idx in [5, 6, 7]:  # Header rows + first data row
        cell = ws.cell(row=row_idx, column=first_run_col_idx)
        template_run_col_style[row_idx] = cell._style
    
    # --- 3) Write master marks (rows) ---
    data_start_row = 7
    
    for idx, mark in enumerate(master_marks):
        row_idx = data_start_row + idx
        
        # Column A: Label
        cell_a = ws.cell(row=row_idx, column=1)
        cell_a.value = mark.get("label", "")
        cell_a._style = template_data_row_styles.get(1, cell_a._style)
        
        # Column B: Required Value (use instrument as fallback)
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.value = mark.get("required_value") or mark.get("instrument", "")
        cell_b._style = template_data_row_styles.get(2, cell_b._style)
        
        # Column C: Tol Min
        cell_c = ws.cell(row=row_idx, column=3)
        cell_c.value = mark.get("tol_min", "")
        cell_c._style = template_data_row_styles.get(3, cell_c._style)
        
        # Column D: Tol Max
        cell_d = ws.cell(row=row_idx, column=4)
        cell_d.value = mark.get("tol_max", "")
        cell_d._style = template_data_row_styles.get(4, cell_d._style)
    
    # --- 4) Write runs (columns) ---
    for run_idx, run in enumerate(runs):
        col_idx = first_run_col_idx + run_idx  # E, F, G, ...
        col_letter = get_column_letter(col_idx)
        
        # Row 5: "Inspected by: <name>"
        cell_5 = ws.cell(row=5, column=col_idx)
        cell_5.value = f"Inspected by: {run.inspected_by}"
        cell_5._style = template_run_col_style.get(5, cell_5._style)
        
        # Row 6: "Inspected at: <timestamp>"
        cell_6 = ws.cell(row=6, column=col_idx)
        cell_6.value = f"Inspected at: {run.inspected_at}"
        cell_6._style = template_run_col_style.get(6, cell_6._style)
        
        # Data rows: fill observed values
        for idx, mark in enumerate(master_marks):
            row_idx = data_start_row + idx
            master_mark_id = mark.get("mark_id", "")
            
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = run.values.get(master_mark_id, "")
            cell._style = template_run_col_style.get(7, cell._style)
            
            # Optional: align center
            if cell.alignment:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=cell.alignment.wrap_text,
                )
    
    # --- 5) Auto-adjust column widths ---
    for col_idx in range(1, first_run_col_idx + len(runs) + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx < first_run_col_idx:
            # Fixed widths for Label/Required/Tol columns
            ws.column_dimensions[col_letter].width = 15
        else:
            # Run columns
            ws.column_dimensions[col_letter].width = 20
    
    # --- 6) Save to bytes ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info(
        f"Generated master report: {len(master_marks)} marks × {len(runs)} runs"
    )
    
    return output.getvalue()