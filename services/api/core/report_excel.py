# services/api/core/report_excel.py
from __future__ import annotations
import io
import os
from typing import Dict, List, Any, Optional
from datetime import datetime
from tempfile import NamedTemporaryFile
from zoneinfo import ZoneInfo
from copy import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


from core.report_pdf import _require_pdfium, _fetch_pdf_bytes

from collections import defaultdict
import gc
from PIL import Image, ImageDraw, ImageFont
from settings import get_settings


def _bytes_to_tempfile(data: bytes, suffix: str = ".png") -> str:
    f = NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        f.write(data)
        f.flush()
        return f.name
    finally:
        f.close()


def _write_merged(ws: Worksheet, coord: str, value) -> None:
    """
    Write `value` to `coord`. If `coord` lies inside a merged range,
    write to that range's top-left cell (required by openpyxl).
    """
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        # Find the merged range that contains this coordinate
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                top_left = ws.cell(row=rng.min_row, column=rng.min_col)
                top_left.value = value
                return
        # If we somehow didn't find the range, fail loudly (shouldn't happen)
        raise RuntimeError(f"Cell {coord} is merged but its range wasn't found.")
    else:
        cell.value = value


def _crop_from_page_image(
    *,
    pil_page,
    rect_norm,
    padding_pct: float,
):
    """
    Crop a normalized rectangle (nx, ny, nw, nh) from a pre-rendered
    PIL page image, with padding. This is essentially the same math
    as _render_crop but WITHOUT re-rendering or drawing boxes.
    """
    from PIL import Image as PilImage  # local import to avoid global dependency

    W, H = pil_page.size
    nx, ny, nw, nh = rect_norm

    rx = round(nx * W)
    ry = round(ny * H)
    rw = round(nw * W)
    rh = round(nh * H)

    pad = round(padding_pct * max(rw, rh))

    x0 = max(0, rx - pad)
    y0 = max(0, ry - pad)
    x1 = min(W, rx + rw + pad)
    y1 = min(H, ry + rh + pad)

    crop = pil_page.crop((x0, y0, x1, y1))
    return crop.convert("RGB")

def _is_filled_value(v: Optional[str]) -> bool:
    """
    A mark is considered 'filled' if observed value is:
      - not empty
      - not 'NA' (case-insensitive)
    """
    if v is None:
        return False
    s = str(v).strip()
    if not s:
        return False
    return s.upper() != "NA"


def _map_get(d: Optional[Dict[str, Any]], key: Any) -> Any:
    """
    Safe getter for dicts where mark_id might have whitespace issues.
    Tries raw key, str(key), stripped str(key).
    """
    if not d or key is None:
        return None

    try:
        if key in d:
            return d.get(key)
    except Exception:
        pass

    s = str(key)
    if s in d:
        return d.get(s)

    ss = s.strip()
    if ss in d:
        return d.get(ss)

    return None


def _excel_hyperlink_formula(url: str, display_text: str) -> str:
    u = (url or "").replace('"', '""')
    t = (display_text or "").replace('"', '""')
    return f'=HYPERLINK("{u}", "{t}")'


def _setup_right_side_area(ws: Worksheet) -> None:
    """
    Make columns K..P usable for page images.
    K is main anchor column.
    """
    # Make K wide; keep rest narrow (image will overflow across columns anyway)
    ws.column_dimensions["K"].width = 160
    for col in ["L", "M", "N", "O", "P"]:
        ws.column_dimensions[col].width = 3


def _draw_marks_on_page_image(
    page_img: Image.Image,
    marks_on_page: List[Dict[str, Any]],
) -> Image.Image:
    """
    Draw editor-like overlay:
      - green rectangle around mark region
      - yellow circular balloon label (A/B/C/...) with a small leader line
    """
    img = page_img.convert("RGB")
    draw = ImageDraw.Draw(img)
    W, H = img.size

    def _idx_to_letters(i: int) -> str:
        # 0 -> A, 1 -> B, ... 25 -> Z, 26 -> AA ...
        i = int(i)
        s = ""
        while True:
            i, rem = divmod(i, 26)
            s = chr(ord("A") + rem) + s
            if i == 0:
                break
            i -= 1
        return s

    def _load_bold_font(size: int) -> ImageFont.ImageFont:
        # Try common DejaVu paths (Linux) and local fallback
        candidates = [
            "DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
        ]
        for p in candidates:
            try:
                return ImageFont.truetype(p, size=size)
            except Exception:
                continue
        # last resort
        return ImageFont.load_default()

    # Slightly thinner boxes, but still visible after Excel scaling
    base = min(W, H)
    rect_stroke = max(2, int(base * 0.0020))      # slightly thinner box
    leader_stroke = max(2, int(base * 0.0020))
    circle_stroke = max(3, int(base * 0.0032))    # thicker circle border


    green = (0, 180, 0)
    black = (0, 0, 0)
    yellow = (255, 235, 59)  # strong yellow

    # Sort marks so labels are stable (A,B,C...) for the page
    marks_sorted = sorted(
        marks_on_page,
        key=lambda m: (int(m.get("order_index", 0)), str(m.get("name", ""))),
    )

    for idx, m in enumerate(marks_sorted):
        try:
            nx = float(m["nx"])
            ny = float(m["ny"])
            nw = float(m["nw"])
            nh = float(m["nh"])
        except Exception:
            continue

        # rectangle coords (pixel)
        rx = int(nx * W)
        ry = int(ny * H)
        rw = int(nw * W)
        rh = int(nh * H)

        x0, y0 = rx, ry
        x1, y1 = rx + rw, ry + rh

        # draw green rectangle (slightly thinner)
        draw.rectangle([x0, y0, x1, y1], outline=green, width=rect_stroke)

        # --- Label text ---
        # Prefer stored label if it's already A/B/C style, else use A,B,C.. by order
        raw_label = (m.get("label") or "").strip()
        if raw_label and len(raw_label) <= 3:
            label = raw_label
        else:
            label = _idx_to_letters(idx)

        # --- Bubble sizing (bigger + editor-like) ---
        # Tie bubble size to mark size so it stays visible even after resize
        #r = max(18, min(40, int(min(rw, rh) * 0.45)))  # smaller bubble
        r = 12
        pad = 6


        # Candidate positions like editor: outside top-left, else top-right, else mid-left
        cx = x0 - (r + pad)
        cy = y0 - (r + pad)

        if cx - r < 0:
            cx = x1 + (r + pad)
            cy = y0 - (r + pad)

        if cy - r < 0:
            cx = x0 - (r + pad)
            cy = y0 + int(rh * 0.2)

        # Clamp inside page
        cx = max(r + 2, min(cx, W - r - 2))
        cy = max(r + 2, min(cy, H - r - 2))

        # Leader line target corner nearest to circle
        target_x = x0 if cx < (x0 + x1) / 2 else x1
        target_y = y0 if cy < (y0 + y1) / 2 else y1
        draw.line([(cx, cy), (target_x, target_y)], fill=black, width=leader_stroke)

        # Draw circle
        draw.ellipse(
            [cx - r, cy - r, cx + r, cy + r],
            fill=yellow,
            outline=black,
            width=circle_stroke,
        )

        # --- Font sizing: fit inside the circle ---
        font_size = max(14, int(r * 1.45))  # bold + readable
        font = _load_bold_font(font_size)

        # shrink until it fits
        for _ in range(10):
            bbox = draw.textbbox((0, 0), label, font=font)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
            if tw <= int(1.45 * r) and th <= int(1.20 * r):
                break
            font_size = max(10, int(font_size * 0.88))
            font = _load_bold_font(font_size)

        # Center text
        bbox = draw.textbbox((0, 0), label, font=font)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        tx = cx - tw / 2
        ty = cy - th / 2

        # draw text with stroke so it stays readable after downscale
        draw.text(
            (tx, ty),
            label,
            fill=black,
            font=font,
            stroke_width=max(1, int(r * 0.04)),  # reduce stroke so bold font shows
            stroke_fill=(255, 255, 255),
        )


    return img


async def generate_report_excel(
    *,
    pdf_url: str,
    marks: List[Dict[str, Any]],
    entries: Dict[str, str],
    user_email: Optional[str],
    mark_set_id: str,
    mark_set_label: str = "",
    part_number: str = "",
    external_id: str = "",
    report_title: str = "",
    dwg_num: str = "",                     # optional drawing number
    padding_pct: float = 0.25,
    render_zoom: float = 2.2,
    logo_url: str = "https://res.cloudinary.com/dbwg6zz3l/image/upload/v1753101276/Black_Blue_ctiycp.png",
    statuses: Optional[Dict[str, str]] = None,  # per-mark status map
) -> bytes:
    """
    Build Excel from template with:
      - Header: Part Number, ID (external_id), MarkSet Name, Created By/At, Drawing No
      - One row per mark:
          A: Label
          B: Thumbnail image of mark region (under "Required Value" header)
          C/D: Tolerance Min/Max (left empty)
          E: Observed Value (user input)
          F: Instrument (from mark["instrument"])
          G: Status (text + colour based on PASS/FAIL/DOUBT – NO dropdown)
          H: Comment (left empty)

    Optimised for:
      - Per-page PDF rendering (render each page once, crop many marks)
      - Using temp files for images so openpyxl never sees raw PIL Images
      - Basic mark cap from settings.max_marks_per_report (default 300)
    """

    # Normalize statuses map
    statuses = statuses or {}

    # ---------- PDF + template ----------
    _require_pdfium()
    pdf_bytes = await _fetch_pdf_bytes(pdf_url)

    template_path = os.path.join(
        os.path.dirname(__file__),
        "../templates/report_template.xlsx",
    )
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active
    # Use row 9 as the "template" for data row styling (borders, fonts, etc.)
    template_row_index = 9
    template_row_cells = {
        col: ws.cell(row=template_row_index, column=col)
        # A..I (1..9) → includes Comment column as well
        for col in range(1, 9 + 1)
    }


    def _apply_row_style(target_row: int) -> None:
        """
        Clone the border/fill/font/alignment/number_format from the template
        data row into the given target_row.
        """
        for col, tmpl_cell in template_row_cells.items():
            cell = ws.cell(row=target_row, column=col)
            cell.border = copy(tmpl_cell.border)
            cell.font = copy(tmpl_cell.font)
            cell.fill = copy(tmpl_cell.fill)
            cell.alignment = copy(tmpl_cell.alignment)
            cell.number_format = tmpl_cell.number_format

    _tempfiles: List[str] = []

    # ---------- Settings: cap marks ----------
    try:
        settings = get_settings()
        max_marks = getattr(settings, "max_marks_per_report", 300)
    except Exception:
        max_marks = 300

    # Sort and apply cap here (extra safety – caller should already filter)
    marks_sorted = sorted(
        marks,
        key=lambda m: (int(m.get("order_index", 0)), str(m.get("name", ""))),
    )
    if len(marks_sorted) > max_marks:
        marks_sorted = marks_sorted[:max_marks]

    # ---------- Group marks by page ----------
    marks_by_page: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for m in marks_sorted:
        try:
            page_idx = int(m.get("page_index", 0))
        except Exception:
            page_idx = 0
        marks_by_page[page_idx].append(m)

    import pypdfium2 as pdfium  # type: ignore
    doc = pdfium.PdfDocument(pdf_bytes)

    def crop_from_page(page_img: Image.Image, rect_norm, padding: float) -> Image.Image:
        """
        Local helper: crop from a pre-rendered page image using the same
        normalized-rect + padding logic as _render_crop, but WITHOUT re-rendering.
        """
        W, H = page_img.size
        nx, ny, nw, nh = rect_norm

        rx = round(nx * W)
        ry = round(ny * H)
        rw = round(nw * W)
        rh = round(nh * H)

        pad_px = round(padding * max(rw, rh))
        x0 = max(0, rx - pad_px)
        y0 = max(0, ry - pad_px)
        x1 = min(W, rx + rw + pad_px)
        y1 = min(H, ry + rh + pad_px)

        return page_img.crop((x0, y0, x1, y1)).convert("RGB")

    try:
        # =====================================================
        # HEADER (matches Excel template layout)
        # =====================================================

        # Row 3: "Report Title: <title>"
        title_val = (report_title or "").strip()
        if title_val:
            _write_merged(ws, "A3", f"Report Title: {title_val}")
        else:
            _write_merged(ws, "A3", "Report Title:")

        # Row 4: "ID: <external_id / mark_set_id>"
        id_val = (external_id or mark_set_id or "").strip()
        if id_val:
            _write_merged(ws, "A4", f"ID: {id_val}")
        else:
            _write_merged(ws, "A4", "ID:")

        # Row 5: "Part Number: <part_number>"
        part_num = (part_number or "").strip()
        if part_num:
            _write_merged(ws, "A5", f"Part Number: {part_num}")
        else:
            _write_merged(ws, "A5", "Part Number:")

        # Row 6: "Inspection Map Name: <mark_set_label>"
        ms_label_val = (mark_set_label or "").strip()
        if ms_label_val:
            _write_merged(ws, "A6", f"Inspection Map Name: {ms_label_val}")
        else:
            _write_merged(ws, "A6", "Inspection Map Name:")


        # Row 7: "Drawing No: <dwg_num>"
        dwg_val = (dwg_num or "").strip()
        if dwg_val:
            _write_merged(ws, "A7", f"Drawing No: {dwg_val}")
        else:
            _write_merged(ws, "A7", "Drawing No:")

        # Row 5 right: "Created By: <email>" in merged E5:G5
        created_by = (user_email or "viewer_user").strip()
        _write_merged(ws, "E5", f"Created By: {created_by}")

        # Row 6 right: "Created At: <IST timestamp>" in merged E6:G6
        ist_time = datetime.now(ZoneInfo("Asia/Kolkata"))
        created_at_str = ist_time.strftime("%Y-%m-%d %H:%M IST")
        _write_merged(ws, "E6", f"Created At: {created_at_str}")

        # ✅ Template has "#VALUE!" stored in A1 (merged A1:I2). Clear it so it doesn't show near logo.
        _write_merged(ws, "A1", "")
        # =====================================================
        # ---------- LOGO ----------
        if logo_url:
            try:
                logo_bytes = await _fetch_pdf_bytes(logo_url)
                logo_path = _bytes_to_tempfile(logo_bytes, suffix=".png")
                _tempfiles.append(logo_path)

                logo_img = OpenpyxlImage(logo_path)  # pass PATH, not PIL
                logo_img.width = 150
                logo_img.height = 60
                ws.add_image(logo_img, "C1")
                del logo_img
            except Exception as e:
                print(f"Logo failed: {e}")

        # =====================================================
        # TABLE ROWS
        # =====================================================
        start_row = 9
        current_row = start_row

        # Process pages in ascending order
        for page_index in sorted(marks_by_page.keys()):
            marks_on_page = marks_by_page[page_index]
            if not marks_on_page:
                continue

            # Render this page ONCE at the requested zoom
            try:
                page = doc[page_index]
                page_img = page.render(scale=render_zoom).to_pil()
            except Exception as e:
                print(f"Failed to render page {page_index}: {e}")
                page_img = None

            # Process all marks on this page (sorted)
            marks_on_page_sorted = sorted(
                marks_on_page,
                key=lambda m: (int(m.get("order_index", 0)), str(m.get("name", ""))),
            )

            for m in marks_on_page_sorted:
                r = current_row
                current_row += 1

                # Ensure this row has the same borders / style as the template data row
                if r >= template_row_index:
                    _apply_row_style(r)

                mark_id = m.get("mark_id", "")
                nx = float(m["nx"])
                ny = float(m["ny"])
                nw = float(m["nw"])
                nh = float(m["nh"])
                label = (m.get("label") or f"Mark {r - start_row + 1}").strip()
                observed = (entries.get(mark_id, "") or "").strip()
                instrument = (m.get("instrument") or "").strip()
                # ✅ Required Value (prefer final, fallback to OCR)
                required_value = (m.get("required_value_final") or m.get("required_value_ocr") or "")
                required_value = str(required_value).strip()

                # ---------- STATUS TEXT + COLOUR ----------
                raw_status = (statuses.get(mark_id, "") or "").strip().upper()

                if raw_status == "PASS":
                    status_text = "Pass"
                    status_color = "FF9AE096"  # Pass: 154,224,150
                elif raw_status == "FAIL":
                    status_text = "Fail"
                    status_color = "FFFD5F67"  # Fail: rgb(253, 95, 103)
                elif raw_status == "DOUBT":
                    status_text = "Doubt"
                    status_color = "FFE6AC89"  # Doubt: 230,172,137
                else:
                    status_text = ""
                    status_color = None
                # ------------------------------------------

                _write_merged(ws, f"A{r}", label)        # A: Label
                # B: Inspection Reference → image goes here (handled below)

                # ✅ C: Required Value
                _write_merged(ws, f"C{r}", required_value)

                # ✅ F: Observed Value
                _write_merged(ws, f"F{r}", observed)


                # ✅ G: Instrument
                _write_merged(ws, f"G{r}", instrument)

                # ✅ H: Status – write text + apply fill
                status_cell = ws.cell(row=r, column=8)  # col 8 = H
                status_cell.value = status_text

                # IMPORTANT:
                # Template row styling may carry a default fill (often green).
                # If status is empty, force clear fill so it stays blank.
                status_cell.fill = PatternFill()  # clears any copied fill

                if status_color:
                    status_cell.fill = PatternFill(
                        start_color=status_color,
                        end_color=status_color,
                        fill_type="solid",
                    )



                ws.row_dimensions[r].height = 75         # row height for thumbnail

                # Image thumbnail into column B
                if page_img is None:
                    continue  # can't render thumbnail without page

                try:
                    crop_img = crop_from_page(
                        page_img,
                        (nx, ny, nw, nh),
                        padding_pct,
                    )

                    # Write to temp file so openpyxl sees a filename
                    bio = io.BytesIO()
                    crop_img.save(bio, format="PNG")
                    crop_png = bio.getvalue()
                    thumb_path = _bytes_to_tempfile(crop_png, suffix=".png")
                    _tempfiles.append(thumb_path)

                    thumb_img = OpenpyxlImage(thumb_path)
                    img_w_px, img_h_px = crop_img.size

                    # Fit within approx 175x100 px box
                    cell_w_px = 175
                    cell_h_px = 100
                    scale = min(cell_w_px / img_w_px, cell_h_px / img_h_px) * 0.9
                    thumb_img.width = int(img_w_px * scale)
                    thumb_img.height = int(img_h_px * scale)

                    ws.add_image(thumb_img, f"B{r}")

                    del crop_img
                    del thumb_img
                    del bio
                except Exception as e:
                    print(f"Image failed for mark on page {page_index}, row {r}: {e}")
                    continue

            # Drop big page image
            if page_img is not None:
                del page_img
                gc.collect()

        # =====================================================
        # RIGHT SIDE (Template): K1 has "To view complete PDF click:"
        # We will:
        # 1) Put hyperlink in K1
        # 2) Insert ONLY pages where at least 1 mark is filled
        # 3) Draw ONLY filled marks on those pages
        # =====================================================

        # Build filled_marks_by_page (page_index -> marks that are filled)
        filled_marks_by_page: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
        for m in marks_sorted:
            mid = m.get("mark_id")
            observed_val = _map_get(entries, mid)
            if _is_filled_value(observed_val):
                try:
                    pidx = int(m.get("page_index", 0))
                except Exception:
                    pidx = 0
                filled_marks_by_page[pidx].append(m)

        pages_with_filled_marks = sorted(filled_marks_by_page.keys())

        # 1) Hyperlink in K1 (your template text cell)
        try:
            _write_merged(ws, "K1", _excel_hyperlink_formula(pdf_url, "To view complete PDF click"))
            c = ws["K1"]
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.font = Font(color="0000EE", underline="single")
        except Exception as e:
            print(f"Failed to set PDF hyperlink in K1: {e}")

        # 2) Setup right side columns
        _setup_right_side_area(ws)

        # 3) Insert annotated pages starting from K2
        right_row = 2
        TARGET_W = 1100 # standard width for page image

        if pages_with_filled_marks:
            for pidx in pages_with_filled_marks:
                try:
                    page = doc[pidx]
                    # Render right-side pages at higher scale so Excel image stays crisp
                    RIGHT_RENDER_ZOOM = max(render_zoom, 20.0)
                    full_img = page.render(scale=RIGHT_RENDER_ZOOM).to_pil()

                except Exception as e:
                    print(f"Failed to render full page {pidx}: {e}")
                    right_row += 5
                    continue

                try:
                    marks_for_overlay = filled_marks_by_page.get(pidx, [])

                    # ✅ Resize FIRST to final width (keeps quality + text won't vanish)
                    if full_img.size[0] != TARGET_W:
                        ratio = TARGET_W / float(full_img.size[0])
                        target_h = max(50, int(full_img.size[1] * ratio))
                        page_for_excel = full_img.resize((TARGET_W, target_h), resample=Image.LANCZOS)
                    else:
                        page_for_excel = full_img

                    # ✅ Draw overlays on the resized image (final resolution)
                    annotated = _draw_marks_on_page_image(page_for_excel, marks_for_overlay)

                    bio = io.BytesIO()
                    annotated.save(bio, format="PNG")
                    png_bytes = bio.getvalue()

                    page_img_path = _bytes_to_tempfile(png_bytes, suffix=".png")
                    _tempfiles.append(page_img_path)

                    ximg = OpenpyxlImage(page_img_path)
                    ximg.width = annotated.size[0]
                    ximg.height = annotated.size[1]

                    ws.add_image(ximg, f"K{right_row}")

                    rows_needed = max(20, int(ximg.height / 20) + 6)
                    right_row += rows_needed

                    del annotated
                    del ximg
                    del bio
                    if page_for_excel is not full_img:
                        del page_for_excel
                except Exception as e:
                    print(f"Right-side insert failed for page {pidx}: {e}")
                    right_row += 5
                finally:
                    try:
                        del full_img
                    except Exception:
                        pass
                    gc.collect()

        # =====================================================
        # =====================================================
        # STATUS: Dropdown + Conditional Formatting (so colour changes when user edits)
        # =====================================================
        last_row = current_row - 1
        if last_row >= start_row:
            status_range = f"H{start_row}:H{last_row}"

            # Dropdown values (allow blank)
            dv = DataValidation(
                type="list",
                formula1='"Pass,Fail,Doubt"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid Status",
                error="Select only from: Pass, Fail, Doubt",
            )
            ws.add_data_validation(dv)
            dv.add(status_range)

            # Conditional formatting fills (matches your RGBs)
            pass_fill = PatternFill(start_color="FF9AE096", end_color="FF9AE096", fill_type="solid")
            fail_fill = PatternFill(start_color="FFFD5F67", end_color="FFFD5F67", fill_type="solid")
            doubt_fill = PatternFill(start_color="FFE6AC89", end_color="FFE6AC89", fill_type="solid")

            # Use first cell formula; Excel will apply relative rows across the range
            ws.conditional_formatting.add(
                status_range,
                FormulaRule(formula=[f'H{start_row}="Pass"'], fill=pass_fill),
            )
            ws.conditional_formatting.add(
                status_range,
                FormulaRule(formula=[f'H{start_row}="Fail"'], fill=fail_fill),
            )
            ws.conditional_formatting.add(
                status_range,
                FormulaRule(formula=[f'H{start_row}="Doubt"'], fill=doubt_fill),
            )

        # =====================================================
        # FINALIZE: save to bytes
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        gc.collect()
        return out.read()

    finally:
        # Cleanup temp files
        for p in _tempfiles:
            try:
                os.unlink(p)
            except Exception:
                pass
